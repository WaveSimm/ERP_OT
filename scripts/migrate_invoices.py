"""
발주처별 인보이스 엑셀 → overseas_orders + contracts 마이그레이션
시트: 총괄, 송금완료, 무환통관  (Sheet1 = 총괄 사본 → 스킵)
"""

import openpyxl
import psycopg2
import re
import sys
import io
import json
from datetime import datetime, date
from decimal import Decimal

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

DB_URL = "host=localhost port=5432 dbname=erp_ot user=erp_user password=erp_password"

wb = openpyxl.load_workbook('References/발주처별 인보이스_260330.xlsx', data_only=True)

def gen_cuid():
    """Generate a cuid-like ID"""
    import random, string, time
    ts = hex(int(time.time() * 1000))[2:]
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=12))
    return f"cm{ts}{rand}"

def parse_date(val):
    """Parse date from cell value, return None if not a valid date"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    # Try common patterns
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # If it contains text like "4월초", "선적지연" etc, return None
    return None

def parse_amount(val):
    """Parse numeric amount"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(',', '').replace('$', '').replace('€', '').replace('£', '')
    try:
        return Decimal(s)
    except:
        return None

def clean_str(val):
    """Clean string value"""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def detect_currency(val):
    """Detect currency from cell value"""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ('EUR', 'EU'):
        return 'EUR'
    if s in ('GBP', 'GB'):
        return 'GBP'
    if s in ('USD', 'US'):
        return 'USD'
    if s in ('KRW', 'KR'):
        return 'KRW'
    return None

def determine_status(row_data, is_completed=False, is_duty_free=False):
    """Determine order status based on available dates"""
    if is_completed:
        return 'CLOSED'
    if is_duty_free:
        if row_data.get('customs_date') or row_data.get('arrival_date'):
            return 'CLOSED'
        return 'CUSTOMS'

    if row_data.get('arrival_date'):
        return 'ARRIVED'
    if row_data.get('customs_date'):
        return 'CUSTOMS'
    if row_data.get('actual_ship_date'):
        return 'SHIPPED'
    if row_data.get('estimated_ship_date') or row_data.get('order_date'):
        return 'ORDERED'
    return 'DRAFT'

def parse_general_row(ws, r, current_currency):
    """Parse a row from 총괄 or 송금완료 sheet"""
    a = clean_str(ws.cell(row=r, column=1).value)  # 계약#
    b = clean_str(ws.cell(row=r, column=2).value)  # 거래처
    c = clean_str(ws.cell(row=r, column=3).value)  # 제작사
    d = clean_str(ws.cell(row=r, column=4).value)  # 품목
    e = clean_str(ws.cell(row=r, column=5).value)  # 통화
    f = ws.cell(row=r, column=6).value              # 발주금액
    g = ws.cell(row=r, column=7).value              # 발주일
    h = ws.cell(row=r, column=8).value              # 선적예정
    i = ws.cell(row=r, column=9).value              # 선적일
    j = ws.cell(row=r, column=10).value             # 통관
    k = ws.cell(row=r, column=11).value             # 입고
    l = clean_str(ws.cell(row=r, column=12).value)  # 입고위치
    m = clean_str(ws.cell(row=r, column=13).value)  # 발주담당자
    n = clean_str(ws.cell(row=r, column=14).value)  # invoice#
    o = ws.cell(row=r, column=15).value              # due date
    p = clean_str(ws.cell(row=r, column=16).value)  # OA#
    q = clean_str(ws.cell(row=r, column=17).value)  # 통관담당
    rr = clean_str(ws.cell(row=r, column=18).value) # 비고

    # Skip currency separator rows
    currency = detect_currency(e)
    if d and str(d).strip() in ('EUR', 'GBP', 'USD', 'JPY'):
        return None, currency or current_currency
    if not (b or c or d):
        return None, current_currency

    # Use detected currency or fallback to current section's currency
    if not currency:
        currency = current_currency

    # Build notes from non-date text in date fields + 비고
    notes_parts = []
    h_date = parse_date(h)
    if h and not h_date:
        notes_parts.append(f"선적예정: {str(h).strip()}")
    j_date = parse_date(j)
    if j and not j_date:
        notes_parts.append(f"통관: {str(j).strip()}")
    o_date = parse_date(o)
    if o and not o_date:
        notes_parts.append(f"due date: {str(o).strip()}")
    if rr:
        notes_parts.append(rr)
    notes = '\n'.join(notes_parts) if notes_parts else None

    amount = parse_amount(f)
    if amount is None:
        amount = Decimal('0')

    return {
        'contract_no': a,
        'client': b,
        'manufacturer': c,
        'item_name': d,
        'currency': currency or 'EUR',
        'amount': amount,
        'order_date': parse_date(g),
        'estimated_ship_date': h_date,
        'actual_ship_date': parse_date(i),
        'customs_date': j_date,
        'arrival_date': parse_date(k),
        'arrival_location': l,
        'ordered_by': m or '미지정',
        'invoice_no': n,
        'due_date': o_date,
        'oa_number': p,
        'customs_handler': q,
        'notes': notes,
    }, current_currency

def parse_duty_free_row(ws, r, current_currency):
    """Parse a row from 무환통관 sheet"""
    a = clean_str(ws.cell(row=r, column=1).value)   # 계약#
    b = clean_str(ws.cell(row=r, column=2).value)   # 거래처
    c = clean_str(ws.cell(row=r, column=3).value)   # 제작사
    d = clean_str(ws.cell(row=r, column=4).value)   # 품목
    e = clean_str(ws.cell(row=r, column=5).value)   # 통화
    f = ws.cell(row=r, column=6).value               # 통관금액
    g = ws.cell(row=r, column=7).value               # 통관일
    h = ws.cell(row=r, column=8).value               # 입고일
    i = clean_str(ws.cell(row=r, column=9).value)   # 입고위치
    j = clean_str(ws.cell(row=r, column=10).value)  # 담당자
    k = clean_str(ws.cell(row=r, column=11).value)  # 통관담당
    l = clean_str(ws.cell(row=r, column=12).value)  # 비고

    # Check columns M, N, O for extra notes (some rows have data there)
    m_val = clean_str(ws.cell(row=r, column=13).value)
    n_val = clean_str(ws.cell(row=r, column=14).value)
    o_val = clean_str(ws.cell(row=r, column=15).value)

    currency = detect_currency(e)
    if d and str(d).strip() in ('EUR', 'GBP', 'USD', 'JPY', ' 품목'):
        return None, currency or current_currency
    if b and str(b).strip() == '거래처':
        return None, current_currency
    if not (b or c or d):
        return None, current_currency

    if not currency:
        currency = current_currency

    amount = parse_amount(f)
    if amount is None:
        amount = Decimal('0')

    notes_parts = []
    if l:
        notes_parts.append(l)
    if m_val:
        notes_parts.append(m_val)
    if o_val:
        notes_parts.append(str(o_val))
    notes = '\n'.join(notes_parts) if notes_parts else None

    return {
        'contract_no': a,
        'client': b,
        'manufacturer': c,
        'item_name': d,
        'currency': currency or 'EUR',
        'amount': amount,
        'order_date': None,
        'estimated_ship_date': None,
        'actual_ship_date': None,
        'customs_date': parse_date(g),
        'arrival_date': parse_date(h),
        'arrival_location': i,
        'ordered_by': j or k or '미지정',
        'invoice_no': None,
        'due_date': None,
        'oa_number': None,
        'customs_handler': k,
        'notes': notes,
    }, current_currency


def main():
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # Check existing data
    cur.execute("SELECT COUNT(*) FROM equipment.overseas_orders")
    existing = cur.fetchone()[0]
    print(f"기존 overseas_orders 수: {existing}")

    # Track created contracts to avoid duplicates
    contract_cache = {}  # contract_number -> id

    # Load existing contracts
    cur.execute("SELECT id, contract_number FROM equipment.contracts")
    for row in cur.fetchall():
        contract_cache[row[1]] = row[0]
    print(f"기존 contracts 수: {len(contract_cache)}")

    # Track existing orders by (contract_no, manufacturer, item_name) to avoid duplicates
    existing_orders = set()
    cur.execute("""
        SELECT o.order_number, o.manufacturer, oi.name
        FROM equipment.overseas_orders o
        LEFT JOIN equipment.overseas_order_items oi ON oi.order_id = o.id
    """)
    for row in cur.fetchall():
        existing_orders.add((row[0], row[1], row[2]))

    order_counter = existing  # for generating order numbers
    stats = {'created': 0, 'skipped': 0, 'errors': 0, 'contracts_created': 0}

    def get_or_create_contract(data):
        """Get existing or create new contract"""
        contract_no = data.get('contract_no')

        if contract_no and contract_no in contract_cache:
            return contract_cache[contract_no]

        # Generate contract number if missing
        if not contract_no:
            # Use manufacturer + client as pseudo contract
            pseudo = f"INV-{data['manufacturer'] or 'UNKNOWN'}-{data['client'] or 'UNKNOWN'}"[:50]
            if pseudo in contract_cache:
                return contract_cache[pseudo]
            contract_no = pseudo

        cid = gen_cuid()
        client = data.get('client') or '미지정'
        manufacturer = data.get('manufacturer') or '미지정'
        item_name = data.get('item_name') or ''

        # Determine contract type
        contract_type = '외자'

        cur.execute("""
            INSERT INTO equipment.contracts (id, contract_number, name, client, manufacturer, category, contract_type, contract_date, status, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'ACTIVE', NOW(), NOW())
            ON CONFLICT (contract_number) DO UPDATE SET updated_at = NOW()
            RETURNING id
        """, (cid, contract_no, item_name[:200] if item_name else f"{manufacturer} 발주", client, manufacturer, '물품', contract_type, data.get('order_date')))

        result_id = cur.fetchone()[0]
        contract_cache[contract_no] = result_id
        stats['contracts_created'] += 1
        return result_id

    def create_order(data, order_type, is_completed=False):
        nonlocal order_counter
        order_counter += 1

        contract_id = get_or_create_contract(data)
        status = determine_status(data, is_completed, order_type == 'DUTY_FREE')

        oid = gen_cuid()
        order_number = f"ORD-{order_counter:05d}"

        # Create the order
        cur.execute("""
            INSERT INTO equipment.overseas_orders
            (id, contract_id, order_number, manufacturer, currency, order_type, status,
             order_date, estimated_ship_date, actual_ship_date, customs_date, arrival_date,
             production_progress, arrival_location, ordered_by, customs_handler,
             invoice_no, due_date, oa_number, total_amount, notes, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s::equipment."OrderCurrency", %s::equipment."OrderType", %s::equipment."OrderStatus",
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, NOW(), NOW())
        """, (
            oid, contract_id, order_number,
            data.get('manufacturer') or '미지정',
            data['currency'],
            order_type,
            status,
            data.get('order_date'),
            data.get('estimated_ship_date'),
            data.get('actual_ship_date'),
            data.get('customs_date'),
            data.get('arrival_date'),
            0,
            data.get('arrival_location'),
            data.get('ordered_by') or '미지정',
            data.get('customs_handler'),
            data.get('invoice_no'),
            data.get('due_date'),
            data.get('oa_number'),
            float(data['amount']),
            data.get('notes'),
        ))

        # Create order item
        item_id = gen_cuid()
        cur.execute("""
            INSERT INTO equipment.overseas_order_items
            (id, order_id, name, quantity, unit_price, amount, receipt_status, created_at, updated_at)
            VALUES (%s, %s, %s, 1, %s, %s, %s, NOW(), NOW())
        """, (
            item_id, oid,
            data.get('item_name') or '품목 미상',
            float(data['amount']),
            float(data['amount']),
            'FULLY_RECEIVED' if status in ('ARRIVED', 'CLOSED') else 'PENDING',
        ))

        stats['created'] += 1

    # ─── 1. 총괄 시트 (현재 진행 중) ───
    print("\n=== 총괄 시트 처리 ===")
    ws = wb['총괄']
    current_currency = 'EUR'
    for r in range(6, ws.max_row + 1):  # 5행=헤더, 6행부터 데이터
        row_data, current_currency = parse_general_row(ws, r, current_currency)
        if row_data:
            try:
                create_order(row_data, 'PURCHASE', is_completed=False)
            except Exception as e:
                print(f"  Row {r} ERROR: {e}")
                stats['errors'] += 1
    print(f"  총괄: {stats['created']}건 생성")

    # ─── 2. 송금완료 시트 (완료 건) ───
    print("\n=== 송금완료 시트 처리 ===")
    prev_created = stats['created']
    ws = wb['송금완료 (2025-현재 )']
    current_currency = 'EUR'
    for r in range(1, ws.max_row + 1):
        row_data, current_currency = parse_general_row(ws, r, current_currency)
        if row_data:
            try:
                create_order(row_data, 'PURCHASE', is_completed=True)
            except Exception as e:
                print(f"  Row {r} ERROR: {e}")
                stats['errors'] += 1
    print(f"  송금완료: {stats['created'] - prev_created}건 생성")

    # ─── 3. 무환통관 시트 ───
    print("\n=== 무환통관 시트 처리 ===")
    prev_created = stats['created']
    ws = wb['무환통관']
    current_currency = 'EUR'
    for r in range(6, ws.max_row + 1):  # 6행=헤더, 7행부터 데이터
        row_data, current_currency = parse_duty_free_row(ws, r, current_currency)
        if row_data:
            try:
                create_order(row_data, 'DUTY_FREE', is_completed=True)
            except Exception as e:
                print(f"  Row {r} ERROR: {e}")
                stats['errors'] += 1
    print(f"  무환통관: {stats['created'] - prev_created}건 생성")

    conn.commit()

    # Final count
    cur.execute("SELECT COUNT(*) FROM equipment.overseas_orders")
    final = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM equipment.contracts")
    final_contracts = cur.fetchone()[0]

    print(f"\n=== 결과 ===")
    print(f"생성: {stats['created']}건 / 에러: {stats['errors']}건")
    print(f"계약 생성: {stats['contracts_created']}건")
    print(f"최종 overseas_orders: {final}건")
    print(f"최종 contracts: {final_contracts}건")

    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
