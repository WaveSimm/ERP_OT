"""Planner import 후 progress 누락 보완.
work log 없는 task에 system work log 1건 추가 후 segment progressPercent 다시 update."""
import sys, io, datetime as dt
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
import openpyxl, requests

EMAIL, PASSWORD = "dev@oceant.com", "dev1234"
API = "http://localhost:3003/api/v1"

# (project_id, excel_path)
TARGETS = [
    ("cmolmp0dr00ejzjt28yyztzxd", "References/planner/[기술팀] 선박-온바다호-2026.xlsx"),
    ("cmolmw4we00tszjt2jnu8099r", "References/planner/[기술팀] KHOA 해양관측부이 차세대 데이터로거 물품 제작.xlsx"),
    ("cmon0oxuf01gq4kz9o1pl5xsh", "References/planner/[사업1팀] KHOA 2026년 해양관측부이 유지관리.xlsx"),
]

def login():
    r = requests.post("http://localhost:3001/api/v1/auth/login",
                      json={"email": EMAIL, "password": PASSWORD}, timeout=10)
    return r.json()["accessToken"]

def H(t): return {"Authorization": f"Bearer {t}", "Content-Type": "application/json"}

def to_iso(d):
    if isinstance(d, dt.datetime): return d.date().isoformat()
    if isinstance(d, dt.date): return d.isoformat()
    return None

def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(9, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(10, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        if row.get("이름"):
            rows.append(row)
    return rows

def fix_project(project_id, excel_path, token):
    print(f"\n📂 Project {project_id} ← {excel_path}")
    excel_rows = parse_excel(excel_path)

    # name → expected progress (이름 매칭 by Excel)
    name_to_progress = {}
    name_to_start = {}
    for row in excel_rows:
        name = str(row.get("이름") or "").strip()
        if row.get("마일스톤") == "예":  # 마일스톤은 task 아님
            continue
        pct = row.get("% 완료")
        if pct is not None:
            progress = round(float(pct) * 100) if float(pct) <= 1 else round(float(pct))
            name_to_progress[name] = max(0, min(100, progress))
        s = to_iso(row.get("시작"))
        if s: name_to_start[name] = s

    # 프로젝트의 모든 task
    r = requests.get(f"{API}/projects/{project_id}/tasks", headers=H(token), timeout=15)
    tasks = r.json()
    print(f"  Tasks: {len(tasks)}")

    # work log 없는 task 식별 + system worklog 추가
    added_wl = 0
    for t in tasks:
        wlr = requests.get(f"{API}/tasks/{t['id']}/work-logs?limit=1",
                           headers=H(token), timeout=10)
        wls = wlr.json() if wlr.ok else {"items": []}
        items = wls if isinstance(wls, list) else wls.get("items", [])
        if not items:
            # Excel 시작일 또는 today
            worked_at = name_to_start.get(t.get("name"), dt.date.today().isoformat())
            wlc = requests.post(f"{API}/tasks/{t['id']}/work-logs", headers=H(token),
                                json={"content": "[system] Planner import 시 자동 생성. 100% 도달을 위한 작업일지 자리.",
                                      "workedAt": worked_at}, timeout=10)
            if wlc.ok:
                added_wl += 1
            else:
                print(f"  ⚠ wl add failed for {t.get('name')}: {wlc.status_code} {wlc.text[:100]}")
    print(f"  + system work logs added: {added_wl}")

    # progress update 재시도
    fixed = 0
    failed = 0
    for t in tasks:
        name = t.get("name")
        if name not in name_to_progress: continue
        target = name_to_progress[name]
        if target == 0: continue
        # 첫 segment id 찾기
        segs = t.get("segments") or []
        if not segs: continue
        seg_id = segs[0]["id"]
        if segs[0].get("progressPercent") == target: continue  # 이미 맞음

        pr = requests.patch(f"{API}/projects/{project_id}/tasks/{t['id']}/segments/{seg_id}",
                            headers=H(token),
                            json={"progressPercent": target, "changeReason": "import progress fix"},
                            timeout=10)
        if pr.ok:
            fixed += 1
        else:
            failed += 1
            print(f"  ⚠ {name}: {pr.status_code} {pr.text[:120]}")
    print(f"  ✓ progress fixed: {fixed}, failed: {failed}")

if __name__ == "__main__":
    token = login()
    for pid, path in TARGETS:
        fix_project(pid, path, token)
    print("\n✅ Done")
