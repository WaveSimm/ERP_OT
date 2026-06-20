"""
Planner Excel → ERP 1회성 import.
8 매핑 결정 적용:
  D1 버킷 무시 / D2 비고+메모 → WorkLog 별도 row
  D3 사용자 자동 reverse / D4/6 우선순위·시간 무시
  D5 체크리스트 → Comment / D7 종속 자동 / D8 마일스톤 → Milestone
"""
import os, sys, io, re, time, datetime as dt
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True)

import openpyxl
import requests
import glob

AUTH_URL = "http://localhost:3001/api/v1/auth/login"
API = "http://localhost:3003/api/v1"
EMAIL = os.environ.get("DEV_EMAIL", "dev@oceant.com")
PASSWORD = os.environ["DEV_PASSWORD"]  # V-22: 평문 제거(env 필수)

# References/planner/ 의 모든 .xlsx 를 자동 인식. 새 팀 플랜은 파일을 넣기만 하면 됨.
#   파일명의 [팀명]·프로젝트명으로 배정을 파싱하므로 '[팀명] 프로젝트.xlsx' 규칙을 지킬 것.
FILES = sorted(glob.glob("References/planner/*.xlsx"))

# ─── HTTP helpers (Session 사용 — keep-alive로 connection 재사용, ~80배 빠름) ──

SESSION = requests.Session()
adapter = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20, max_retries=0)
SESSION.mount("http://", adapter)

def login():
    # v1.6.4 (2026-05-16): cookie 기반 인증으로 전환. accessToken 대신 cookie가 SESSION에 자동 보관.
    r = SESSION.post(AUTH_URL, json={"email": EMAIL, "password": PASSWORD}, timeout=10)
    r.raise_for_status()
    # 토큰 반환 의미 없지만 호환성 위해 더미값
    return "cookie-auth"

def H(token):
    # cookie 기반 — Authorization 헤더 불필요. Content-Type만.
    return {"Content-Type": "application/json"}

def post(path, token, body):
    r = SESSION.post(f"{API}{path}", headers=H(token), json=body, timeout=15)
    if not r.ok:
        raise RuntimeError(f"POST {path} {r.status_code} {r.text[:200]}")
    return r.json() if r.text else None

def patch(path, token, body):
    r = SESSION.patch(f"{API}{path}", headers=H(token), json=body, timeout=15)
    if not r.ok:
        raise RuntimeError(f"PATCH {path} {r.status_code} {r.text[:200]}")
    return r.json() if r.text else None

def put(path, token, body):
    r = SESSION.put(f"{API}{path}", headers=H(token), json=body, timeout=15)
    if not r.ok:
        raise RuntimeError(f"PUT {path} {r.status_code} {r.text[:200]}")
    return r.json() if r.text else None

def get(path, token, params=None):
    r = SESSION.get(f"{API}{path}", headers=H(token), params=params or {}, timeout=15)
    if not r.ok:
        raise RuntimeError(f"GET {path} {r.status_code} {r.text[:200]}")
    return r.json()

# ─── 사용자(auth_user) 조회 ─────────────────────────────────────────────────
# v1.6.4 (2026-05-16): 자원-모델-분리(사람=auth_users 단일) 적용
# 기존 /resources 폐기 → /users 사용. resource.userId(email) 매핑 제거.

_auth_users_cache = None
def all_auth_users(token):
    global _auth_users_cache
    if _auth_users_cache is None:
        r = SESSION.get("http://localhost:3001/api/v1/users",
                        headers=H(token), timeout=15)
        if r.ok:
            d = r.json()
            _auth_users_cache = d.get("items", d if isinstance(d, list) else [])
        else:
            _auth_users_cache = []
    return _auth_users_cache

# 호환성 alias — 기존 코드의 all_resources 호출을 그대로 살림
def all_resources(token):
    return all_auth_users(token)

def reverse_name(planner_name: str) -> str:
    """'재엽 김' → '김재엽'. 한 단어면 그대로."""
    parts = planner_name.strip().split()
    if len(parts) == 2:
        return parts[1] + parts[0]
    return planner_name.strip()

def resolve_resource_id(planner_name: str, token) -> str | None:
    target = reverse_name(planner_name)
    for u in all_auth_users(token):
        if u.get("name") == target:
            return u["id"]
    return None

def resolve_owner_id(planner_name: str, token) -> str | None:
    """Planner 이름 → auth_user.id. resource_id와 동일."""
    return resolve_resource_id(planner_name, token)

# ─── Excel 파싱 ──────────────────────────────────────────────────────────────

def to_iso_date(d) -> str | None:
    if isinstance(d, dt.datetime):
        return d.date().isoformat()
    if isinstance(d, dt.date):
        return d.isoformat()
    if isinstance(d, str) and d.strip():
        return d.strip()[:10]
    return None

def parse_excel(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # meta
    meta = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, 8) if ws.cell(r, 1).value}
    # tasks
    headers = [ws.cell(9, c).value for c in range(1, ws.max_column + 1)]
    tasks = []
    for r in range(10, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        if not row.get("이름"):
            continue
        tasks.append(row)
    return meta, tasks

# ─── 의존성 파싱 ─────────────────────────────────────────────────────────────

DEP_RE = re.compile(r"^(\d+)(FS|SS|FF|SF)?(?:[+\-]\d+\s*일?)?$", re.IGNORECASE)

def parse_dependencies(s) -> list[tuple[int, str]]:
    """'3,4FS' 또는 '3FF' 등 → [(taskNumber, type), ...]"""
    if not s: return []
    deps = []
    for token in str(s).split(","):
        token = token.strip()
        m = DEP_RE.match(token)
        if m:
            deps.append((int(m.group(1)), (m.group(2) or "FS").upper()))
    return deps

# ─── WBS parent 결정 ─────────────────────────────────────────────────────────

def parent_overview(overview: str) -> str | None:
    """1.2.3 → 1.2, 1.2 → 1, 1 → None"""
    if not overview: return None
    parts = str(overview).split(".")
    if len(parts) <= 1: return None
    return ".".join(parts[:-1])

# ─── 단일 파일 import ────────────────────────────────────────────────────────

def import_file(path: str, token):
    print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"📂 {path}")
    print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    meta, raw_tasks = parse_excel(path)
    proj_name = meta.get("프로젝트 이름", "Imported Project")
    owner_planner_name = meta.get("플랜 소유자")  # 예: "윤송 심" (Planner Excel B3)
    print(f"프로젝트: {proj_name}")
    print(f"태스크: {len(raw_tasks)}개")

    # 1) Project 생성 — owner 매핑 (D9: 플랜 소유자 → Project.ownerId = auth_user.id)
    body = {"name": proj_name, "description": f"Planner import — {dt.date.today()}"}
    if owner_planner_name:
        owner_id = resolve_owner_id(owner_planner_name, token)  # auth_user.id
        if owner_id:
            body["ownerId"] = owner_id
            print(f"  소유자: {owner_planner_name} → {reverse_name(owner_planner_name)} (auth_user {owner_id})")
        else:
            print(f"  ⚠ 소유자 매핑 실패: {owner_planner_name} (요청자가 owner 됨)")
    proj = post("/projects", token, body)
    project_id = proj["id"]
    print(f"  ✓ Project created: {project_id}")

    # 2) Task 생성 Pass 1 (parentId 없이) — 시점 task(isMilestone=true) 통합
    overview_to_task_id = {}
    task_number_to_task_id = {}
    task_records = []  # 후속 처리용

    # 사전 분석: 자식이 있는 overview 집합 (마일스톤이라도 자식이 있으면 일반 task로 강제)
    all_ovs = {str(r.get("개요 번호") or "").strip() for r in raw_tasks if r.get("개요 번호")}
    has_children_set = {ov for ov in all_ovs if any(o != ov and o.startswith(ov + ".") for o in all_ovs)}

    for row in raw_tasks:
        task_num = row.get("작업 번호")
        overview = str(row.get("개요 번호") or "").strip()
        is_milestone_raw = (row.get("마일스톤") == "예")
        # 자식이 있는 task는 시점 task일 수 없음 → 일반 task로 강제 + 경고
        if is_milestone_raw and overview in has_children_set:
            print(f"  ⚠ 마일스톤 '{row.get('이름')}' (개요 {overview})은 자식이 있어 일반 task로 변환")
            is_milestone = False
        else:
            is_milestone = is_milestone_raw

        # sortOrder = 작업번호 (A 컬럼) — Excel의 트리 순서 그대로 보존
        sort_order = task_num if isinstance(task_num, int) else 0
        body = {
            "name": str(row.get("이름") or "(이름없음)"),
            "sortOrder": sort_order,
            "isMilestone": is_milestone,  # 시점 task 플래그
        }
        t = post(f"/projects/{project_id}/tasks", token, body)
        overview_to_task_id[overview] = t["id"]
        task_number_to_task_id[task_num] = t["id"]
        task_records.append({
            "row": row,
            "is_milestone": is_milestone,
            "task_id": t["id"],
            "overview": overview,
            "task_num": task_num,
        })
    ms_cnt = len([r for r in task_records if r['is_milestone']])
    print(f"  ✓ Tasks created: {len(task_records)} (시점 {ms_cnt})")

    # 3) parentId 설정 (Pass 2 — overview 기반)
    parent_set = 0
    for rec in task_records:
        ov = rec["overview"]
        pov = parent_overview(ov)
        if pov and pov in overview_to_task_id:
            patch(f"/projects/{project_id}/tasks/{rec['task_id']}", token,
                  {"parentId": overview_to_task_id[pov]})
            parent_set += 1
    print(f"  ✓ parentId set: {parent_set}")

    # 4) Segment 생성 — 시점 task는 단일 segment(start=end)
    for rec in task_records:
        row = rec["row"]
        start = to_iso_date(row.get("시작"))
        end = to_iso_date(row.get("마침"))
        if not start: continue
        # 시점 task: end = start
        if rec["is_milestone"]:
            end = start
        elif not end or end < start:
            end = start
        seg = post(f"/projects/{project_id}/tasks/{rec['task_id']}/segments", token, {
            "name": "시점" if rec["is_milestone"] else str(row.get("이름") or "구간"),
            "startDate": start,
            "endDate": end,
        })
        rec["segment_id"] = seg["id"]
    seg_count = len([r for r in task_records if r.get('segment_id')])
    print(f"  ✓ Segments created: {seg_count}")

    # 5) Assignment 생성 (D3 사용자 매핑) — 시점 task도 자원 배정 허용 (OQ-1)
    asn_count = 0
    for rec in task_records:
        if not rec.get("segment_id"): continue
        names = rec["row"].get("할당 대상")
        if not names: continue
        for n in str(names).split(","):
            rid = resolve_resource_id(n.strip(), token)
            if not rid:
                print(f"  ⚠ unmapped name: {n.strip()!r}")
                continue
            try:
                put(f"/projects/{project_id}/tasks/{rec['task_id']}/segments/{rec['segment_id']}/assignments",
                    token, {"resourceId": rid, "allocationMode": "PERCENT", "allocationPercent": 100})
                asn_count += 1
            except Exception as e:
                print(f"  ⚠ assignment failed for {n}: {e}")
    print(f"  ✓ Assignments: {asn_count}")

    # 6) Milestone 폐기 — Task isMilestone=true로 통합 ("마일스톤-시점태스크-회귀" PDCA)

    # 7) WorkLog 생성 (D2 — 비고 + 메모 → 각각 별도 row, 시점 task 포함)
    wl_count = 0
    for rec in task_records:
        if not rec.get("task_id"): continue
        row = rec["row"]
        start_iso = to_iso_date(row.get("시작")) or dt.date.today().isoformat()

        # 비고 (J)
        bigo = row.get("비고")
        if bigo and str(bigo).strip():
            try:
                post(f"/tasks/{rec['task_id']}/work-logs", token, {
                    "content": f"[비고] {str(bigo).strip()}",
                    "workedAt": start_iso,
                })
                wl_count += 1
            except Exception as e:
                print(f"  ⚠ worklog 비고 failed: {e}")

        # 메모 (R)
        memo = row.get("메모")
        if memo and str(memo).strip():
            try:
                post(f"/tasks/{rec['task_id']}/work-logs", token, {
                    "content": f"[작업일지] {str(memo).strip()}",
                    "workedAt": start_iso,
                })
                wl_count += 1
            except Exception as e:
                print(f"  ⚠ worklog 메모 failed: {e}")

        # 체크리스트 (D5 — Comment 로 통합)
        # 백엔드 comment API는 이번 PDCA 범위 밖, 임시로 work log 1건에 합침
        chk = row.get("체크리스트 항목")
        if chk and str(chk).strip():
            items = str(chk).strip().split(";")
            content = "[체크리스트]\n" + "\n".join(f"☐ {x.strip()}" for x in items if x.strip())
            try:
                post(f"/tasks/{rec['task_id']}/work-logs", token, {
                    "content": content,
                    "workedAt": start_iso,
                })
                wl_count += 1
            except Exception as e:
                print(f"  ⚠ worklog 체크리스트 failed: {e}")
    print(f"  ✓ WorkLogs: {wl_count}")

    # 8) Dependency 생성 (D7 — Task↔Task만)
    dep_count = 0
    for rec in task_records:
        if not rec.get("task_id"): continue
        deps = parse_dependencies(rec["row"].get("종속 대상"))
        for pred_num, dep_type in deps:
            pred_id = task_number_to_task_id.get(pred_num)
            if not pred_id:
                print(f"  ⚠ dep predecessor not found: task#{pred_num}")
                continue
            try:
                post(f"/projects/{project_id}/dependencies", token, {
                    "predecessorTaskId": pred_id,
                    "successorTaskId": rec["task_id"],
                    "dependencyType": dep_type,
                    "lag": 0,
                })
                dep_count += 1
            except Exception as e:
                print(f"  ⚠ dep failed (#{pred_num} {dep_type}): {e}")
    print(f"  ✓ Dependencies: {dep_count}")

    # 9) Segment progressPercent 업데이트 (시점 task 포함)
    prog_count = 0
    for rec in task_records:
        if not rec.get("segment_id"): continue
        row = rec["row"]
        pct = row.get("% 완료")
        if pct is None: continue
        # 0~1 비율을 0~100으로
        progress = round(float(pct) * 100) if float(pct) <= 1 else round(float(pct))
        progress = max(0, min(100, progress))
        if progress == 0: continue
        try:
            patch(f"/projects/{project_id}/tasks/{rec['task_id']}/segments/{rec['segment_id']}", token, {
                "progressPercent": progress,
                "changeReason": "Planner import",
            })
            prog_count += 1
        except Exception as e:
            print(f"  ⚠ progress update failed: {e}")
    print(f"  ✓ Progress updates: {prog_count}")

    print(f"\n  📊 Project {project_id} import complete")
    return project_id

# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # CLI args 우선, 없으면 기본 FILES 사용
    targets = sys.argv[1:] if len(sys.argv) > 1 else FILES

    print("Login...")
    token = login()
    print(f"OK")

    print("\nResource lookup...")
    rs = all_resources(token)
    print(f"  active resources: {len(rs)}")

    created = []
    for path in targets:
        try:
            pid = import_file(path, token)
            created.append((path, pid))
        except Exception as e:
            print(f"\n❌ FAILED for {path}: {e}")
            raise

    print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"✅ Imported {len(created)} project(s):")
    for path, pid in created:
        print(f"  - {pid} ← {path}")
