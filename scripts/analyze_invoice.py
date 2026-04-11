import openpyxl
import json
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('References/발주처별 인보이스_260330.xlsx', data_only=True)
out = []

out.append(f"=== 시트 목록: {wb.sheetnames} ===")
out.append(f"총 시트 수: {len(wb.sheetnames)}\n")

for sname in wb.sheetnames:
    ws = wb[sname]
    out.append(f"\n{'='*60}")
    out.append(f"=== 시트: {sname} | 행: {ws.max_row} | 열: {ws.max_column} ===")

    # Print first 30 rows to understand structure
    row_limit = min(50, ws.max_row + 1)
    for r in range(1, row_limit):
        vals = []
        for c in range(1, min(30, ws.max_column + 1)):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                val_str = str(cell.value)[:80]
                vals.append((cell.coordinate, val_str))
        if vals:
            out.append(f"Row {r}: {json.dumps(vals, ensure_ascii=False)}")
        elif r <= 5:
            out.append(f"Row {r}: (empty)")

with open('scripts/analyze_invoice_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print("Done")
