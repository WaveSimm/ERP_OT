import openpyxl
import json

wb = openpyxl.load_workbook('References/2025년 원가정산서.xlsx', data_only=True)
out = []

# Find settlements with multiple items across all sheets
for sname in wb.sheetnames[:-1]:
    ws = wb[sname]
    boundaries = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and "수 입 원 가" in str(v):
            boundaries.append(r)

    for i, start in enumerate(boundaries):
        end = boundaries[i+1] if i+1 < len(boundaries) else ws.max_row

        # Find section 6 items
        items_found = 0
        for r in range(start, end):
            a_val = ws.cell(row=r, column=1).value  # A column (inventory no)
            b_val = ws.cell(row=r, column=2).value   # B column (product name)
            x_val = ws.cell(row=r, column=24).value   # X column (quantity)
            ar_val = ws.cell(row=r, column=44).value   # AR column (KRW unit price)

            if a_val and str(a_val).startswith('E') and x_val:
                items_found += 1

        if items_found >= 2:
            out.append(f"\n=== {sname} 정산서 {i+1} (행 {start}~{end-1}) - {items_found}개 품목 ===")
            # Print section 6 area
            for r in range(start, end):
                b_val = ws.cell(row=r, column=2).value
                if b_val and "6. 모델" in str(b_val):
                    # Print from here to end of block
                    for r2 in range(r, min(r + items_found + 5, end)):
                        vals = []
                        for c in range(1, 56):
                            cell = ws.cell(row=r2, column=c)
                            if cell.value is not None:
                                vals.append((cell.coordinate, str(cell.value)[:60]))
                        if vals:
                            out.append(f"Row {r2}: {json.dumps(vals, ensure_ascii=False)}")
                    break

# Also check: are there settlements with multiple remittance rows? (already seen one with 2)
out.append("\n\n=== 송금 행수 분포 ===")
remittance_counts = {}
for sname in wb.sheetnames[:-1]:
    ws = wb[sname]
    boundaries = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and "수 입 원 가" in str(v):
            boundaries.append(r)

    for i, start in enumerate(boundaries):
        end = boundaries[i+1] if i+1 < len(boundaries) else ws.max_row
        rem_count = 0
        in_remittance = False
        for r in range(start, end):
            b_val = ws.cell(row=r, column=2).value
            if b_val and "(1) 송금액" in str(b_val):
                in_remittance = True
                rem_count += 1
                continue
            if in_remittance:
                if b_val and "소" in str(b_val):
                    break
                n_val = ws.cell(row=r, column=14).value  # N column
                if n_val is not None and n_val != '' and n_val != ' ':
                    rem_count += 1

        remittance_counts[rem_count] = remittance_counts.get(rem_count, 0) + 1

out.append(f"송금 행수별 정산서 수: {json.dumps(remittance_counts)}")

# Check section 5 values (수입원가합계)
out.append("\n=== Section 5 값 샘플 (수입원가, 공급가액, 부가세) ===")
count = 0
for sname in wb.sheetnames[:-1]:
    ws = wb[sname]
    for r in range(1, ws.max_row + 1):
        b_val = ws.cell(row=r, column=2).value
        if b_val and "금액" == str(b_val).strip() and ws.cell(row=r-1, column=2).value and "구분" in str(ws.cell(row=r-1, column=2).value):
            h_val = ws.cell(row=r, column=8).value    # 수입원가
            n_val = ws.cell(row=r, column=14).value   # 공급가액
            x_val = ws.cell(row=r, column=24).value   # 부가세
            ae_val = ws.cell(row=r, column=31).value   # 비고
            if h_val:
                out.append(f"{sname} Row {r}: 수입원가={h_val}, 공급가액={n_val}, 부가세={x_val}, 비고={str(ae_val)[:60] if ae_val else ''}")
                count += 1
                if count >= 10:
                    break
    if count >= 10:
        break

# Check currency - is it always EUR? Look at section 6 headers
out.append("\n=== 통화 확인 (Section 6 헤더) ===")
currencies = set()
for sname in wb.sheetnames[:3]:
    ws = wb[sname]
    for r in range(1, ws.max_row + 1):
        b_val = ws.cell(row=r, column=2).value
        if b_val and "구분" == str(b_val).strip():
            ai_val = ws.cell(row=r, column=35).value  # AI column
            ao_val = ws.cell(row=r, column=41).value  # AO column
            if ai_val and "(" in str(ai_val):
                currencies.add(str(ai_val).strip())
                out.append(f"{sname} Row {r}: AI={ai_val}, AO={ao_val}")

out.append(f"통화 종류: {currencies}")

with open('scripts/analyze_cost_multi_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print("Done")
