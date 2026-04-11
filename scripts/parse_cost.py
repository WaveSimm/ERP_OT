import openpyxl
import json

wb = openpyxl.load_workbook('References/2025년 원가정산서.xlsx', data_only=True)
out = []

ws = wb[wb.sheetnames[0]]  # 1월
out.append("=== 1월 rows 26-46 (section 5,6) ===")
for r in range(26, 46):
    vals = []
    for c in range(1, 68):
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            vals.append((cell.coordinate, str(cell.value)))
    if vals:
        out.append(json.dumps(vals, ensure_ascii=False))

# Count settlements per month
out.append("\n=== Settlements per month ===")
for sname in wb.sheetnames[:-1]:  # skip template
    ws2 = wb[sname]
    count = 0
    for r in range(1, ws2.max_row + 1):
        v = ws2.cell(row=r, column=2).value
        if v and "수 입 원 가" in str(v):
            count += 1
    out.append(f"{sname}: {count} settlements")

with open('scripts/parse_cost_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print("Done")
