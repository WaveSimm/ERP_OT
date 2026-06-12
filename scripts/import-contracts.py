# 계약 마이그레이션 — References/contracts/YYYY년 계약파일리스트.xlsx → contracts 정규화 JSON
#  - R2가 헤더(R0/R1은 메모). R3부터 데이터.
#  - 연도마다 컬럼 위치가 달라 '헤더명 기반' 매핑.
#  - contractNumber = #YY-{A열 연번}  (예 #25-1). 기존 DB와 동일 형식.
# 실행: python scripts/import-contracts.py  →  tmp/contracts_import.json (CONTRACTS_OUT 로 변경 가능)
import openpyxl, glob, os, json, re
from datetime import datetime, date

SRC = sorted(glob.glob("References/contracts/*계약파일리스트*.xlsx"))
OUT = os.environ.get("CONTRACTS_OUT", "tmp/contracts_import.json")  # 리눅스/윈도 공통 상대경로

def norm(s):
    return re.sub(r"\s+", "", str(s or "")).strip()

def find_header_row(ws):
    # '거래처'와 '계약건명/품명'이 같이 있는 행을 헤더로 (보통 R2 = index 2)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
        cells = [norm(c) for c in row]
        if any("거래처" in c for c in cells) and any(("계약건명" in c or "품명" in c) for c in cells):
            return i, row
    return 2, None

def col_map(header):
    """헤더 셀 배열 → 필드별 컬럼 인덱스."""
    cells = [norm(c) for c in header]
    m = {}
    for i, c in enumerate(cells):
        if not c: continue
        if "거래처" in c and "client" not in m: m["client"] = i
        elif c == "담당" and "clientContact" not in m: m["clientContact"] = i
        elif "제작사" in c and "manufacturer" not in m: m["manufacturer"] = i
        elif ("계약건명" in c or "품명" in c) and "name" not in m: m["name"] = i
        elif c == "구분" and "category" not in m: m["category"] = i
        elif "계약종류" in c and "contractType" not in m: m["contractType"] = i
        elif "계약일자" in c and "contractDate" not in m: m["contractDate"] = i
        elif c == "납기" and "deadline" not in m: m["deadline"] = i
        elif (c.startswith("계약담당") or c == "담당자") and "manager" not in m: m["manager"] = i
        elif c == "비고" and "notes" not in m: m["notes"] = i
    return m

def as_date(v):
    if isinstance(v, (datetime, date)): return v.strftime("%Y-%m-%d")
    return None

def cat(v):
    s = norm(v)
    if "용역" in s: return "용역"
    return "물품"  # 상품/물품/기타 → 물품

def ctype(v):
    s = norm(v)
    if "외자" in s: return "외자"
    return "내자"

records = []
seen = set()
per_year = {}
for f in SRC:
    base = os.path.basename(f)
    yy = base[2:4]  # 'YYYY년...' → YY
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    hidx, header = find_header_row(ws)
    if header is None:
        for r in ws.iter_rows(min_row=hidx+1, max_row=hidx+1, values_only=True): header = r
    m = col_map(header)
    cnt = 0
    for row in ws.iter_rows(min_row=hidx+2, values_only=True):
        def cell(key):
            i = m.get(key)
            return row[i] if (i is not None and i < len(row)) else None
        seq = row[0] if len(row) > 0 else None
        name = cell("name")
        if not (name and norm(name)): continue
        if not isinstance(seq, (int, float)): continue
        cnum = "#%s-%d" % (yy, int(seq))
        if cnum in seen: continue  # 파일 내 중복 연번 방지
        seen.add(cnum)
        records.append({
            "contractNumber": cnum,
            "name": str(name).strip()[:200],
            "client": (str(cell("client")).strip()[:200] if cell("client") else ""),
            "clientContact": (str(cell("clientContact")).strip()[:100] if cell("clientContact") else None),
            "manufacturer": (str(cell("manufacturer")).strip()[:200] if cell("manufacturer") else None),
            "category": cat(cell("category")),
            "contractType": ctype(cell("contractType")),
            "contractDate": as_date(cell("contractDate")),
            "deadline": as_date(cell("deadline")),
            "manager": (str(cell("manager")).strip()[:100] if cell("manager") else None),
            "notes": (str(cell("notes")).strip()[:500] if cell("notes") else None),
            "year": "20" + yy,
        })
        cnt += 1
    per_year[base[:5]] = cnt
    wb.close()

os.makedirs(os.path.dirname(OUT) or ".", exist_ok=True)
json.dump(records, open(OUT, "w", encoding="utf-8"), ensure_ascii=False)
print("연도별:", per_year)
print("총 레코드:", len(records))
print("샘플:", json.dumps(records[0], ensure_ascii=False)[:300] if records else "none")
print("→", OUT)
