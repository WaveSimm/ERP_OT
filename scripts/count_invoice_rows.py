import openpyxl
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('References/발주처별 인보이스_260330.xlsx', data_only=True)

for sname in wb.sheetnames:
    ws = wb[sname]
    # Count rows that have data in column B, C, or D (거래처/제작사/품목)
    data_rows = 0
    last_row = 0
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value  # 거래처
        c = ws.cell(row=r, column=3).value  # 제작사
        d = ws.cell(row=r, column=4).value  # 품목
        if b or c or d:
            # Skip header rows and currency separator rows
            if d and str(d).strip() in ('EUR', 'GBP', 'USD', 'JPY', '품목', ' 품목'):
                continue
            if b and str(b).strip() in ('거래처',):
                continue
            data_rows += 1
            last_row = r
    print(f"{sname}: max_row={ws.max_row}, data_rows={data_rows}, last_data_row={last_row}")
