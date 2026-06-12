"""기존 import된 프로젝트의 task sortOrder를 Excel 작업번호로 보정."""
import sys, io, datetime as dt
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
import openpyxl, requests

API = "http://localhost:3003/api/v1"
EMAIL, PASSWORD = "dev@oceant.com", "dev1234"

TARGETS = [
    ("cmolmp0dr00ejzjt28yyztzxd", "References/planner/[기술팀] 선박-온바다호-2026.xlsx"),
    ("cmolmw4we00tszjt2jnu8099r", "References/planner/[기술팀] KHOA 해양관측부이 차세대 데이터로거 물품 제작.xlsx"),
]

def login():
    r = requests.post("http://localhost:3001/api/v1/auth/login",
                      json={"email": EMAIL, "password": PASSWORD}, timeout=10)
    return r.json()["accessToken"]

def H(t): return {"Authorization": f"Bearer {t}", "Content-Type": "application/json"}

def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(9, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(10, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        if row.get("이름"):
            rows.append(row)
    return rows

def parent_overview(overview):
    if not overview: return None
    parts = str(overview).split(".")
    if len(parts) <= 1: return None
    return ".".join(parts[:-1])

def fix(project_id, excel_path, token):
    print(f"\n📂 {project_id} ← {excel_path}")
    rows = parse(excel_path)
    # name → sortOrder (작업번호) — 시점 task 포함 (마일스톤-시점태스크-회귀 PDCA 후 통합)
    name_to_sort = {}
    overview_to_task_name = {}
    name_to_parent_overview = {}
    for row in rows:
        name = str(row.get("이름") or "").strip()
        num = row.get("작업 번호")
        ov = str(row.get("개요 번호") or "").strip()
        if not name or not isinstance(num, int): continue
        name_to_sort[name] = num
        overview_to_task_name[ov] = name
        name_to_parent_overview[name] = parent_overview(ov)

    # ERP tasks 조회 (시점 task 포함)
    r = requests.get(f"{API}/projects/{project_id}/tasks", headers=H(token), timeout=15)
    tasks = r.json()
    print(f"  Tasks: {len(tasks)}")

    fixed = 0
    skipped = 0
    name_to_task_id = {t.get("name"): t["id"] for t in tasks}
    for t in tasks:
        name = t.get("name")
        target = name_to_sort.get(name)
        if target is None: continue
        body = {}
        if t.get("sortOrder") != target:
            body["sortOrder"] = target
        # 시점 task가 parent 누락된 경우 parent 보정
        if t.get("isMilestone") and not t.get("parentId"):
            pov = name_to_parent_overview.get(name)
            parent_task_name = overview_to_task_name.get(pov) if pov else None
            parent_task_id = name_to_task_id.get(parent_task_name) if parent_task_name else None
            if parent_task_id:
                body["parentId"] = parent_task_id
        if not body:
            skipped += 1
            continue
        pr = requests.patch(f"{API}/projects/{project_id}/tasks/{t['id']}",
                            headers=H(token), json=body, timeout=10)
        if pr.ok:
            fixed += 1
        else:
            print(f"  ⚠ patch failed: {name}: {pr.status_code} {pr.text[:80]}")
    print(f"  ✓ tasks fixed: {fixed}, already-correct: {skipped}")

if __name__ == "__main__":
    token = login()
    for pid, path in TARGETS:
        fix(pid, path, token)
    print("\n✅ Done")
