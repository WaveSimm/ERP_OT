"""
2025년 수입원가정산서 엑셀 → DB 마이그레이션 스크립트
Usage: python scripts/migrate_cost_settlements.py
"""

import openpyxl
import psycopg2
import re
import sys
import io
from datetime import datetime
from decimal import Decimal

# Fix Windows console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

DB_URL = "host=localhost port=5432 dbname=erp_ot user=erp_user password=erp_password"
EXCEL_PATH = "References/2025년 원가정산서.xlsx"
CREATED_BY = "migration"  # 마이그레이션 유저

# Column indices (1-based)
COL_A = 1    # 재고번호 (items)
COL_B = 2    # 텍스트 labels
COL_H = 8    # 신고번호, 공급자, 신고일, 수입원가
COL_N = 14   # 외화금액, 관세금액, 공급가액
COL_X = 24   # 환율, 부가세, 수량
COL_AE = 31  # 원화금액, 비고, 외화단가
COL_AI = 35  # 통화 (section 6 header)
COL_AL = 38  # Invoice, 외화금액 (items)
COL_AR = 44  # 원화단가 (items)
COL_AX = 50  # 원화총액 (items)

DUTY_MAP = {
    "1)관세": "TARIFF",
    "2)운반비(국외)": "OVERSEAS_FREIGHT",
    "3)운반비(국내)": "DOMESTIC_FREIGHT",
    "4)통관수수료": "CUSTOMS_FEE",
    "5)창고보관료": "WAREHOUSE_FEE",
    "6)취급수수료": "HANDLING_FEE",
}


def safe_num(val, default=0):
    """Convert value to float safely."""
    if val is None:
        return default
    try:
        v = float(val)
        return v
    except (ValueError, TypeError):
        return default


def safe_str(val):
    """Convert value to string safely."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_date(val):
    """Parse date from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def extract_currency(ws, header_row):
    """Extract currency from section 6 header (AI column)."""
    ai_val = ws.cell(row=header_row, column=COL_AI).value
    if ai_val:
        s = str(ai_val).strip().strip("()")
        if s in ("USD", "EUR", "GBP", "CAD", "JPY", "CHF", "CNY"):
            return s
    return "USD"


def parse_settlement(ws, start_row, end_row):
    """Parse a single settlement record from the worksheet."""
    record = {
        "declaration_no": None,
        "supplier": None,
        "declaration_date": None,
        "currency": "USD",
        "total_import_cost": 0,
        "supply_amount": 0,
        "vat": 0,
        "sale_info": None,
        "notes": None,
        "remittances": [],
        "duties": [],
        "items": [],
    }

    # Scan rows within this settlement block
    i = start_row
    while i < end_row:
        b_val = safe_str(ws.cell(row=i, column=COL_B).value)

        # 1. 신고번호
        if b_val and "1. 신고번호" in b_val:
            record["declaration_no"] = safe_str(ws.cell(row=i, column=COL_H).value)

        # 2. 공급자
        elif b_val and "2. 공급자" in b_val:
            record["supplier"] = safe_str(ws.cell(row=i, column=COL_H).value)

        # 3. 신고일
        elif b_val and "3. 신고일" in b_val:
            record["declaration_date"] = parse_date(ws.cell(row=i, column=COL_H).value)

        # (1) 송금액 rows
        elif b_val and "(1) 송금액" in b_val:
            # First remittance row
            rem = parse_remittance_row(ws, i)
            if rem:
                record["remittances"].append(rem)
            # Continue checking subsequent rows for additional remittances
            j = i + 1
            while j < end_row:
                bj = safe_str(ws.cell(row=j, column=COL_B).value)
                if bj and ("소" in bj or "(2)" in bj):
                    break
                # Check if this row has remittance data (N column has value)
                n_val = ws.cell(row=j, column=COL_N).value
                h_val = ws.cell(row=j, column=COL_H).value
                if n_val is not None and safe_num(n_val) != 0:
                    rem = parse_remittance_row(ws, j)
                    if rem:
                        record["remittances"].append(rem)
                j += 1
            i = j - 1  # Will be incremented at end of loop

        # (2) 관세,통관,운송료 - duty rows
        elif b_val and b_val.strip().replace(" ", "") in [s.replace(" ", "") for s in ["1)관세", "2)운반비(국외)", "3)운반비(국내)", "4)통관수수료", "5)창고보관료", "6)취급수수료"]]:
            duty_type = None
            clean_b = b_val.strip()
            for k, v in DUTY_MAP.items():
                if k.replace(" ", "") in clean_b.replace(" ", ""):
                    duty_type = v
                    break
            if duty_type:
                amount = safe_num(ws.cell(row=i, column=COL_N).value)
                vat = safe_num(ws.cell(row=i, column=COL_X).value)
                awb = safe_str(ws.cell(row=i, column=COL_AE).value)
                # Also check column A for account name
                a_val = safe_str(ws.cell(row=i, column=COL_A).value)
                if amount > 0 or awb:
                    record["duties"].append({
                        "type": duty_type,
                        "amount": amount,
                        "vat": vat if vat and isinstance(vat, (int, float)) and vat > 0 else 0,
                        "awb_no": awb,
                        "notes": a_val,
                    })

        # 5. 수입원가합계 - next row has values
        elif b_val and "5. 수입원가합계" in b_val:
            # Skip to "금액" row (2 rows below)
            for j in range(i + 1, min(i + 4, end_row)):
                bj = safe_str(ws.cell(row=j, column=COL_B).value)
                if bj and "금액" in bj:
                    record["total_import_cost"] = safe_num(ws.cell(row=j, column=COL_H).value)
                    record["supply_amount"] = safe_num(ws.cell(row=j, column=COL_N).value)
                    record["vat"] = safe_num(ws.cell(row=j, column=COL_X).value)
                    record["sale_info"] = safe_str(ws.cell(row=j, column=COL_AE).value)
                    break

        # 6. 모델,규격 section
        elif b_val and "6. 모델" in b_val:
            # Next row is header, then items
            header_row = i + 1
            record["currency"] = extract_currency(ws, header_row)
            j = header_row + 1
            while j < end_row:
                a_val = safe_str(ws.cell(row=j, column=COL_A).value)
                b_item = safe_str(ws.cell(row=j, column=COL_B).value)
                x_qty = ws.cell(row=j, column=COL_X).value

                # Item row: has inventory number or product name with quantity
                if (a_val and (a_val.startswith("E") or a_val.startswith("S"))) or (b_item and x_qty):
                    if not a_val and not b_item:
                        j += 1
                        continue

                    item = {
                        "inventory_no": a_val,
                        "name": b_item or a_val or "Unknown",
                        "quantity": int(safe_num(x_qty, 1)),
                        "foreign_unit_price": safe_num(ws.cell(row=j, column=COL_AE).value),
                        "foreign_amount": safe_num(ws.cell(row=j, column=COL_AL).value),
                        "unit_price": safe_num(ws.cell(row=j, column=COL_AR).value),
                        "amount": safe_num(ws.cell(row=j, column=COL_AX).value),
                    }
                    record["items"].append(item)
                elif not a_val and not b_item:
                    # Empty row = end of items
                    break
                j += 1

        i += 1

    # If total_import_cost is 0, calculate from remittances + duties
    if record["total_import_cost"] == 0:
        rem_total = sum(r["krw_amount"] for r in record["remittances"])
        duty_total = sum(d["amount"] for d in record["duties"])
        record["total_import_cost"] = rem_total + duty_total

    # If supply_amount is 0, calculate from total_import_cost
    if record["supply_amount"] == 0 and record["total_import_cost"] > 0:
        record["supply_amount"] = round(record["total_import_cost"] / 1.1)
        record["vat"] = round(record["total_import_cost"] - record["supply_amount"])

    return record


def parse_remittance_row(ws, row):
    """Parse a single remittance row."""
    foreign_amount = safe_num(ws.cell(row=row, column=COL_N).value)
    exchange_rate = safe_num(ws.cell(row=row, column=COL_X).value)
    krw_amount = safe_num(ws.cell(row=row, column=COL_AE).value)
    remittance_date = parse_date(ws.cell(row=row, column=COL_H).value)
    invoice_no = safe_str(ws.cell(row=row, column=COL_AL).value)

    if foreign_amount == 0 and krw_amount == 0:
        return None

    return {
        "remittance_date": remittance_date,
        "foreign_amount": foreign_amount,
        "exchange_rate": exchange_rate,
        "krw_amount": krw_amount,
        "invoice_no": invoice_no,
    }


def generate_cuid():
    """Generate a cuid-like ID."""
    import time
    import random
    import string
    ts = hex(int(time.time() * 1000))[2:]
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=16))
    return f"cm{ts}{rand}"


def insert_settlement(cur, record):
    """Insert a settlement record into the database."""
    if not record["declaration_no"]:
        return None

    settlement_id = generate_cuid()

    # Check if declaration_no already exists
    cur.execute(
        "SELECT id FROM equipment.import_cost_settlements WHERE declaration_no = %s",
        (record["declaration_no"],)
    )
    if cur.fetchone():
        print(f"  SKIP (duplicate): {record['declaration_no']}")
        return None

    # Calculate totalExtraCost (sum of duties)
    total_extra = sum(d["amount"] for d in record["duties"])

    cur.execute("""
        INSERT INTO equipment.import_cost_settlements
        (id, declaration_no, supplier, declaration_date, currency,
         total_import_cost, total_extra_cost, supply_amount, vat,
         sale_info, notes, created_by, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
    """, (
        settlement_id,
        record["declaration_no"],
        record["supplier"] or "Unknown",
        record["declaration_date"] or datetime.now().date(),
        record["currency"],
        record["total_import_cost"],
        total_extra,
        record["supply_amount"],
        record["vat"],
        record["sale_info"],
        record["notes"],
        CREATED_BY,
    ))

    # Insert remittances
    for rem in record["remittances"]:
        rem_id = generate_cuid()
        cur.execute("""
            INSERT INTO equipment.cost_remittances
            (id, settlement_id, remittance_date, foreign_amount, exchange_rate, krw_amount, invoice_no)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            rem_id,
            settlement_id,
            rem["remittance_date"] or record["declaration_date"] or datetime.now().date(),
            rem["foreign_amount"],
            rem["exchange_rate"],
            rem["krw_amount"],
            rem["invoice_no"],
        ))

    # Insert duties
    for duty in record["duties"]:
        duty_id = generate_cuid()
        cur.execute("""
            INSERT INTO equipment.cost_duties
            (id, settlement_id, type, amount, vat, awb_no, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            duty_id,
            settlement_id,
            duty["type"],
            duty["amount"],
            duty["vat"],
            duty["awb_no"],
            duty["notes"],
        ))

    # Insert items
    for item in record["items"]:
        item_id = generate_cuid()
        cur.execute("""
            INSERT INTO equipment.cost_items
            (id, settlement_id, inventory_no, name, quantity,
             foreign_unit_price, foreign_amount, unit_price, amount)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            item_id,
            settlement_id,
            item["inventory_no"],
            item["name"],
            item["quantity"],
            item["foreign_unit_price"] if item["foreign_unit_price"] > 0 else None,
            item["foreign_amount"] if item["foreign_amount"] > 0 else None,
            item["unit_price"],
            item["amount"],
        ))

    return settlement_id


def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Connect to DB
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    total_inserted = 0
    total_skipped = 0
    total_errors = 0

    for sname in wb.sheetnames[:-1]:  # Skip template sheet
        ws = wb[sname]
        print(f"\n--- {sname} ---")

        # Find settlement boundaries
        boundaries = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=COL_B).value
            if v and "수 입 원 가" in str(v):
                boundaries.append(r)

        print(f"  Found {len(boundaries)} settlements")

        for i, start in enumerate(boundaries):
            end = boundaries[i + 1] if i + 1 < len(boundaries) else ws.max_row + 1

            try:
                record = parse_settlement(ws, start, end)
                sid = insert_settlement(cur, record)
                if sid:
                    rem_count = len(record["remittances"])
                    duty_count = len(record["duties"])
                    item_count = len(record["items"])
                    print(f"  [{i+1}/{len(boundaries)}] {record['declaration_no']} | "
                          f"{record['supplier']} | {record['currency']} | "
                          f"₩{record['total_import_cost']:,.0f} | "
                          f"송금:{rem_count} 관세:{duty_count} 품목:{item_count}")
                    total_inserted += 1
                else:
                    total_skipped += 1
            except Exception as e:
                print(f"  ERROR [{i+1}] row {start}: {e}")
                total_errors += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"\n{'='*60}")
    print(f"Migration complete!")
    print(f"  Inserted: {total_inserted}")
    print(f"  Skipped (duplicates): {total_skipped}")
    print(f"  Errors: {total_errors}")
    print(f"  Total: {total_inserted + total_skipped + total_errors}")


if __name__ == "__main__":
    main()
