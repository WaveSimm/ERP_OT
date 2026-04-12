import openpyxl
import json

wb = openpyxl.load_workbook('References/2025년 원가정산서.xlsx', data_only=True)
out = []

# Analyze first sheet (1월) thoroughly - find settlement boundaries
ws = wb['1월']
out.append(f"=== 1월 총 행수: {ws.max_row}, 총 열수: {ws.max_column} ===\n")

# Print all rows with data for first 2 settlements
out.append("=== 1월 전체 데이터 (행 1~50) ===")
for r in range(1, 51):
    vals = []
    for c in range(1, 60):
        cell = ws.cell(row=r, column=c)
        if cell.value is not None:
            vals.append((cell.coordinate, str(cell.value)[:80]))
    if vals:
        out.append(f"Row {r}: {json.dumps(vals, ensure_ascii=False)}")
    else:
        out.append(f"Row {r}: (empty)")

# Find all "수 입 원 가" rows to identify settlement boundaries
out.append("\n=== 1월 '수 입 원 가' 위치 ===")
boundaries = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=2).value
    if v and "수 입 원 가" in str(v):
        boundaries.append(r)
        out.append(f"Row {r}: {v}")

out.append(f"\n총 {len(boundaries)}개 정산서, 경계 행: {boundaries}")

# For second settlement, show full structure
if len(boundaries) >= 2:
    start = boundaries[1]
    end = boundaries[2] if len(boundaries) > 2 else ws.max_row
    out.append(f"\n=== 2번째 정산서 (행 {start}~{end-1}) ===")
    for r in range(start, min(end, start + 25)):
        vals = []
        for c in range(1, 60):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                vals.append((cell.coordinate, str(cell.value)[:80]))
        if vals:
            out.append(f"Row {r}: {json.dumps(vals, ensure_ascii=False)}")

# Check a month with more settlements (6월)
ws6 = wb['6월']
out.append(f"\n=== 6월 총 행수: {ws6.max_row} ===")
boundaries6 = []
for r in range(1, ws6.max_row + 1):
    v = ws6.cell(row=r, column=2).value
    if v and "수 입 원 가" in str(v):
        boundaries6.append(r)
out.append(f"6월 정산서 수: {len(boundaries6)}")

# Check for multiple items in one settlement (section 6 with multiple rows)
if len(boundaries6) >= 2:
    # Find one with multiple items
    for i in range(min(5, len(boundaries6))):
        start = boundaries6[i]
        end = boundaries6[i+1] if i+1 < len(boundaries6) else ws6.max_row
        out.append(f"\n=== 6월 정산서 {i+1} (행 {start}~{end-1}) ===")
        for r in range(start, min(end, start + 25)):
            vals = []
            for c in range(1, 60):
                cell = ws6.cell(row=r, column=c)
                if cell.value is not None:
                    vals.append((cell.coordinate, str(cell.value)[:80]))
            if vals:
                out.append(f"Row {r}: {json.dumps(vals, ensure_ascii=False)}")

with open('scripts/analyze_cost_detail_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print("Done")
